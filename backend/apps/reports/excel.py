"""Excel (.xlsx) exports for the generic financial/operational reports.

Reuses the exact per-report flattener logic from ``pdf.py`` (same report keys,
same query params, same shared table structure) and renders it with openpyxl:
a title/subtitle block, a bold frozen header row, right/left aligned data cells,
money cells as real numbers with a '#,##0.00' format, and bold section/total
rows.
"""
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import NotFound
from rest_framework.views import APIView

from .pdf import _report_specs

MONEY_FORMAT = '#,##0.00'
INT_FORMAT = '#,##0'

_HEADER_FILL = PatternFill('solid', fgColor='1F2937')
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(italic=True, size=10, color='555555')
_SECTION_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)


def _as_number(text):
    """Parse a flattener-formatted money/number string back to a number.

    Returns (value, number_format) or (None, None) when it is not a plain
    number (e.g. blanks or percentages like '95.0%')."""
    if text is None:
        return None, None
    raw = str(text).strip()
    if raw == '' or raw.endswith('%'):
        return None, None
    stripped = raw.replace(',', '')
    try:
        value = Decimal(stripped)
    except (InvalidOperation, ValueError):
        return None, None
    fmt = MONEY_FORMAT if ('.' in stripped or ',' in raw) else INT_FORMAT
    return value, fmt


def _filename_suffix(data):
    if data.get('as_of_date'):
        return data['as_of_date']
    if data.get('start') and data.get('end'):
        return f"{data['start']}_{data['end']}"
    if data.get('term'):
        return f"term-{data['term']}"
    return timezone.localdate().isoformat()


def _build_workbook(flat):
    columns = flat['columns']
    ncols = len(columns)
    wb = Workbook()
    ws = wb.active
    ws.title = (flat['title'] or 'Report')[:31]

    # --- Title / subtitle block -------------------------------------------
    ws.cell(row=1, column=1, value=flat['title']).font = _TITLE_FONT
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    subtitle = flat.get('subtitle') or ''
    ws.cell(row=2, column=1, value=subtitle).font = _SUBTITLE_FONT
    if ncols > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    # --- Header row -------------------------------------------------------
    header_row = 3
    for col, spec in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=spec['label'])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal='right' if spec['align'] == 'right' else 'left', vertical='center'
        )
    ws.freeze_panes = f'A{header_row + 1}'

    # --- Data rows --------------------------------------------------------
    row_idx = header_row + 1
    max_len = [len(str(c['label'])) for c in columns]
    for row in flat['rows']:
        style = row.get('style', '')
        cells = row['cells']
        if style == 'section':
            # Section headings carry a single label spanning the sheet width.
            label = cells[0] if cells else ''
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = _SECTION_FONT
            if ncols > 1:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
            max_len[0] = max(max_len[0], len(str(label)))
            row_idx += 1
            continue

        bold = style in ('total', 'bold')
        for col, text in enumerate(cells, start=1):
            if col > ncols:
                break
            align = columns[col - 1]['align']
            cell = ws.cell(row=row_idx, column=col)
            if align == 'right':
                value, fmt = _as_number(text)
                if value is not None:
                    cell.value = value
                    cell.number_format = fmt
                else:
                    cell.value = text
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = text
                cell.alignment = Alignment(horizontal='left')
            if bold:
                cell.font = _TOTAL_FONT
            max_len[col - 1] = max(max_len[col - 1], len(str(text)))
        row_idx += 1

    # --- Column widths ----------------------------------------------------
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len[col - 1] + 2, 10), 50)

    return wb


class ReportExcelView(APIView):
    """Generic Excel export: /api/reports/xlsx/<report_key>/ with the same query
    params as the underlying JSON report view (and as the PDF export)."""

    def get(self, request, report_key):
        specs = _report_specs()
        if report_key not in specs:
            raise NotFound(f'Unknown report: {report_key}')
        view_class, flatten = specs[report_key]
        data = view_class().build(request)
        if isinstance(data, dict) and data.get('error'):
            return JsonResponse({'error': data['error']}, status=400)

        flat = flatten(data)
        wb = _build_workbook(flat)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'{report_key}-{_filename_suffix(data)}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
